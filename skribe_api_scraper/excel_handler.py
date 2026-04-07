import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
from loguru import logger
from utils import normalize_text

class ExcelHandler:
    def __init__(self):
        self.output_cols = [
            "Journalist_ID", "Email", "Email_2", "Phone", "Twitter", "LinkedIn", 
            "Beat", "Media_Types", "City", "State", "Country", "Title", 
            "Profile_URL", "Outlet_Match_Confidence", "Match_Score", 
            "Scrape_Status", "Scrape_Error", "Scraped_At"
        ]

    def read_input(self, file_path):
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Input file not found: {file_path}")
            
        try:
            df = pd.read_excel(file_path, dtype=str)
            df = df.fillna("")
            
            # Detect columns
            name_col = self.detect_column(df, ["journalist", "name", "full name"])
            pub_col = self.detect_column(df, ["publication", "outlet", "media", "agency", "organization"])
            
            if not name_col or not pub_col:
                found = ", ".join(df.columns)
                raise ValueError(f"Could not auto-detect Name/Publication columns. Found: {found}")
            
            # Normalize
            df[name_col] = df[name_col].apply(normalize_text)
            df[pub_col] = df[pub_col].apply(normalize_text)
            
            # Drop empty
            initial_count = len(df)
            df = df[~((df[name_col] == "") & (df[pub_col] == ""))]
            final_count = len(df)
            
            logger.info(f"Loaded {final_count} journalists. Skipped {initial_count - final_count} empty rows.")
            
            return df, name_col, pub_col
        except Exception as e:
            logger.error(f"Failed to read input Excel: {e}")
            raise

    def detect_column(self, df, keywords):
        for col in df.columns:
            if any(k in col.lower() for k in keywords):
                return col
        return None

    def write_output(self, df, output_path, summary_data=None):
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Main Sheet
            df.to_excel(writer, index=False, sheet_name="Enriched Journalists")
            
            # Summary Sheet
            if summary_data:
                summary_df = pd.DataFrame(list(summary_data.items()), columns=["Metric", "Value"])
                summary_df.to_excel(writer, index=False, sheet_name="Run Summary")
        
        # Apply formatting
        self.apply_formatting(output_path)
        logger.info(f"Output saved to {output_path}")

    def apply_formatting(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb["Enriched Journalists"]
        
        header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Status fills
        fills = {
            "SUCCESS": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "PARTIAL": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "NOT_FOUND": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
            "ERROR": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "STOPPED": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        }
        
        # Format headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Freeze top row
        ws.freeze_panes = "A2"
        
        # Row-based formatting for status
        status_col_idx = None
        for i, cell in enumerate(ws[1], 1):
            if cell.value == "Scrape_Status":
                status_col_idx = i
                break
                
        if status_col_idx:
            for row in ws.iter_rows(min_row=2):
                status_val = row[status_col_idx - 1].value
                if status_val in fills:
                    fill = fills[status_val]
                    for cell in row:
                        cell.fill = fill
        
        # Auto-size columns
        for column_cells in ws.columns:
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            length = min(length + 2, 50) # Limit max width
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length
            
        wb.save(file_path)
